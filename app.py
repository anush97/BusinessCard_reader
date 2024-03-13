from flask import Flask, request, render_template, redirect, url_for
import os
import boto3
from botocore.exceptions import NoCredentialsError
import pytesseract
from PIL import Image
import spacy
import pandas as pd
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl import Workbook

app = Flask(__name__)

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
EXCEL_FILE = 'extracted_data.xlsx'

# Ensure upload folder exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def download_model_from_s3():
    s3_bucket_name = 'buscardmodel'  # S3 Bucket name
    s3_model_key = 'model-last/'  # The key prefix for the model files in S3
    local_model_dir = 'model-last'  # Local directory to store the model

    s3 = boto3.client(
        's3',
        aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID'),
        aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY'),
        region_name=os.environ.get('AWS_DEFAULT_REGION')  # Using environment variable for AWS region
    )

    try:
        # List all objects within the specified prefix
        model_files = s3.list_objects(Bucket=s3_bucket_name, Prefix=s3_model_key)['Contents']
        for file in model_files:
            file_key = file['Key']
            # Construct the local file path
            local_file_path = os.path.join(local_model_dir, file_key.replace(s3_model_key, ''))
            
            # Ensure the directory exists
            local_file_dir = os.path.dirname(local_file_path)
            if not os.path.exists(local_file_dir):
                os.makedirs(local_file_dir, exist_ok=True)

            # Download the file
            print(f"Downloading {file_key} to {local_file_path}")
            s3.download_file(s3_bucket_name, file_key, local_file_path)
    except NoCredentialsError:
        print("AWS credentials not available")

# Call the function to start the download when the app starts
download_model_from_s3()

# Load spaCy model
nlp = spacy.load(os.path.join('model-last'))

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_image():
    if 'image' not in request.files:
        return 'No image part', 400
    image = request.files['image']
    if image.filename == '':
        return 'No selected image', 400
    if image and allowed_file(image.filename):
        filename = secure_filename(image.filename)
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        image.save(image_path)
        return redirect(url_for('extract_info', filename=filename))
    else:
        return 'Invalid file type', 400

@app.route('/extract_info')
def extract_info():
    filename = request.args.get('filename')
    if not filename:
        return 'Missing data', 400
    
    image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    extracted_text = pytesseract.image_to_string(Image.open(image_path))

    doc = nlp(extracted_text)
    info = {'NAME': '', 'DES': '', 'ORG': '', 'PHONE': '', 'ADD': '', 'EMAIL': '', 'WEB': ''}
    for ent in doc.ents:
        if info[ent.label_]:
            info[ent.label_] += f", {ent.text}"
        else:
            info[ent.label_] = ent.text

    return render_template('verify.html', info=info)

@app.route('/write_excel', methods=['POST'])
def write_to_excel():
    headers = [
        'Company Name', 'Person Name', 'Designation', 
        'Address', 'Email', 'Mobile no.', 'Website', 
        'Industry (Biscuits, Dairy, Beverages Etc)'
    ]

    form_to_header_mapping = {
        'ORG': 'Company Name', 
        'NAME': 'Person Name', 
        'DES': 'Designation', 
        'ADD': 'Address', 
        'EMAIL': 'Email', 
        'PHONE': 'Mobile no.', 
        'WEB': 'Website'
    }

    info = {header: '' for header in headers}
    for form_field, header in form_to_header_mapping.items():
        info[header] = request.form.get(form_field, '')

    df_new = pd.DataFrame([info.values()])

    try:
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)

    sheet.append(list(info.values()))

    try:
        workbook.save(EXCEL_FILE)
        df = pd.read_excel(EXCEL_FILE)
        excel_data_html = df.to_html(index=False, classes='excel-table')
        return render_template('success.html', excel_table=excel_data_html)
    except Exception as e:
        return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
